"""Structured data ingestion pipeline: Loads ParcelPilot_Assessment_Data.xlsx into SQLite."""

import os
from pathlib import Path
from typing import Dict, Tuple
import openpyxl
from src.config import DB_PATH, RAW_DATA_DIR
from src.db import get_db_connection, init_db


def ingest_structured_data(
    xlsx_path: Path = RAW_DATA_DIR / "ParcelPilot_Assessment_Data.xlsx",
    db_path: Path = DB_PATH
) -> Dict[str, int]:
    """Load Excel sheets into SQLite database tables including dataset snapshot metadata."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Source data file not found at: {xlsx_path}")

    init_db(db_path)
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    conn = get_db_connection(db_path)
    cursor = conn.cursor()

    row_counts: Dict[str, int] = {}

    # 1. Parse README sheet for metadata
    if "README" in wb.sheetnames:
        readme_sheet = wb["README"]
        meta_items = []
        for row in readme_sheet.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            key = str(row[0]).strip().lower().replace(" ", "_")
            val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            meta_items.append((key, val))

        cursor.executemany(
            "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
            meta_items
        )
        row_counts["meta"] = len(meta_items)

    # 2. Ingest accounts sheet
    if "accounts" in wb.sheetnames:
        accounts_sheet = wb["accounts"]
        rows = list(accounts_sheet.iter_rows(values_only=True))
        if rows:
            header = [str(col).strip() for col in rows[0]]
            account_data = []
            for r in rows[1:]:
                if not r or not r[0]:
                    continue
                row_dict = dict(zip(header, r))
                account_data.append((
                    str(row_dict.get("account_id", "")).strip(),
                    str(row_dict.get("account_name", "")).strip(),
                    str(row_dict.get("plan", "")).strip(),
                    str(row_dict.get("status", "")).strip(),
                    str(row_dict.get("csm", "")).strip() if row_dict.get("csm") else None,
                    str(row_dict.get("contract_file", "")).strip() if row_dict.get("contract_file") else None,
                    1 if row_dict.get("premium_support") in (True, 1, "True", "true") else 0,
                    str(row_dict.get("notes", "")).strip() if row_dict.get("notes") else None,
                ))

            cursor.executemany("""
                INSERT OR REPLACE INTO accounts
                (account_id, account_name, plan, status, csm, contract_file, premium_support, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, account_data)
            row_counts["accounts"] = len(account_data)

    # 3. Ingest orders sheet
    if "orders" in wb.sheetnames:
        orders_sheet = wb["orders"]
        rows = list(orders_sheet.iter_rows(values_only=True))
        if rows:
            header = [str(col).strip() for col in rows[0]]
            order_data = []
            for r in rows[1:]:
                if not r or not r[0]:
                    continue
                row_dict = dict(zip(header, r))
                order_data.append((
                    str(row_dict.get("order_id", "")).strip(),
                    str(row_dict.get("account_id", "")).strip(),
                    str(row_dict.get("carrier", "")).strip(),
                    str(row_dict.get("status", "")).strip(),
                    str(row_dict.get("booked_at", "")).strip(),
                    str(row_dict.get("pickup_window_start", "")).strip() if row_dict.get("pickup_window_start") else None,
                    str(row_dict.get("pickup_window_end", "")).strip() if row_dict.get("pickup_window_end") else None,
                    str(row_dict.get("pickup_actual_at", "")).strip() if row_dict.get("pickup_actual_at") else None,
                    float(row_dict.get("shipment_fee_inr", 0.0) or 0.0),
                    1 if row_dict.get("carrier_fault") in (True, 1, "True", "true") else 0,
                    1 if row_dict.get("customer_fault") in (True, 1, "True", "true") else 0,
                    str(row_dict.get("cancellation_requested_at", "")).strip() if row_dict.get("cancellation_requested_at") else None,
                    str(row_dict.get("notes", "")).strip() if row_dict.get("notes") else None,
                ))

            cursor.executemany("""
                INSERT OR REPLACE INTO orders
                (order_id, account_id, carrier, status, booked_at, pickup_window_start, pickup_window_end,
                 pickup_actual_at, shipment_fee_inr, carrier_fault, customer_fault, cancellation_requested_at, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, order_data)
            row_counts["orders"] = len(order_data)

    # 4. Ingest tickets sheet
    if "tickets" in wb.sheetnames:
        tickets_sheet = wb["tickets"]
        rows = list(tickets_sheet.iter_rows(values_only=True))
        if rows:
            header = [str(col).strip() for col in rows[0]]
            ticket_data = []
            for r in rows[1:]:
                if not r or not r[0]:
                    continue
                row_dict = dict(zip(header, r))
                ticket_data.append((
                    str(row_dict.get("ticket_id", "")).strip(),
                    str(row_dict.get("account_id", "")).strip(),
                    str(row_dict.get("created_at", "")).strip(),
                    str(row_dict.get("status", "")).strip(),
                    str(row_dict.get("subject", "")).strip(),
                    str(row_dict.get("description", "")).strip(),
                    str(row_dict.get("channel", "")).strip(),
                    str(row_dict.get("assigned_to", "")).strip() if row_dict.get("assigned_to") else None,
                    str(row_dict.get("last_customer_message_at", "")).strip() if row_dict.get("last_customer_message_at") else None,
                    str(row_dict.get("historical_resolution", "")).strip() if row_dict.get("historical_resolution") else None,
                ))

            cursor.executemany("""
                INSERT OR REPLACE INTO tickets
                (ticket_id, account_id, created_at, status, subject, description, channel,
                 assigned_to, last_customer_message_at, historical_resolution)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, ticket_data)
            row_counts["tickets"] = len(ticket_data)

    conn.commit()
    conn.close()
    return row_counts


if __name__ == "__main__":
    counts = ingest_structured_data()
    print("Ingestion complete:")
    for tbl, count in counts.items():
        print(f"  - Table '{tbl}': {count} rows")
